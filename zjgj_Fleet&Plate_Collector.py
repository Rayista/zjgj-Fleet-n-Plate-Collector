# -*- coding: utf-8 -*-

import json
import os
import sys
import time
import threading
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed

import requests
from openpyxl import Workbook, load_workbook


if getattr(sys, 'frozen', False):
    BASE_DIR = os.path.dirname(os.path.abspath(sys.executable))
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))

LINE_CONFIG_FILE = os.path.join(
    BASE_DIR,
    'lines.json'
)

EXCEL_FILE = os.path.join(
    BASE_DIR,
    'vehicle_id_code.xlsx'
)

INTERVAL_SECONDS = 10 * 60
MAX_WORKERS = 15
REQUEST_TIMEOUT = 10

API_URL = (
    'http://121.40.11.192:8080/lyx/lineBus/getLinePipe'
    '?lineId={}&upOrDownStr={}'
)

HEADERS = [
    '线路号',
    '上下行',
    '自编号',
    '车牌号',
    '发车时间',
    '工号',
    '运营公司'
]

# 定义相互串数据的关联线路组
CROSS_LINE_GROUPS = [
    {'zj_id': '1001',   'rg_id': '971001'},  # 100路 <-> 新K001路
    {'zj_id': '1021',   'rg_id': '971021'},  # 102路 <-> 新K2路
    {'zj_id': '1011',   'rg_id': '971101'},  # 101路 <-> 新K101路
]

# 构建快速查找映射表
LINE_GROUP_MAP = {}
for group in CROSS_LINE_GROUPS:
    LINE_GROUP_MAP[group['zj_id']] = group
    LINE_GROUP_MAP[group['rg_id']] = group

_thread_local = threading.local()


def get_session():
    """获取当前线程的HTTP Session。"""

    if not hasattr(_thread_local, 'session'):
        session = requests.Session()
        session.headers.update({
            'User-Agent': (
                'Mozilla/5.0 '
                '(Windows NT 10.0; Win64; x64) '
                'AppleWebKit/537.36 '
                '(KHTML, like Gecko) '
                'Chrome/120 Safari/537.36'
            )
        })
        _thread_local.session = session

    return _thread_local.session


def load_lines():
    """读取线路配置。"""

    if not os.path.exists(LINE_CONFIG_FILE):
        raise FileNotFoundError(
            f'找不到线路配置文件：{LINE_CONFIG_FILE}'
        )

    with open(
        LINE_CONFIG_FILE,
        'r',
        encoding='utf-8'
    ) as f:
        config = json.load(f)

    if not isinstance(config, dict):
        raise ValueError(
            'lines.json格式错误：最外层必须是对象。'
        )

    lines = config.get('lines')

    if not isinstance(lines, list):
        raise ValueError(
            'lines.json中必须存在lines数组。'
        )

    result = []

    for index, item in enumerate(lines, start=1):

        if not isinstance(item, dict):
            print(
                f'警告：第{index}个线路配置不是对象，已跳过。'
            )
            continue

        line_id = item.get('id')
        line_name = item.get('name')
        company = item.get('company', '')

        if line_id is None or not str(line_id).strip():
            print(
                f'警告：第{index}个线路配置缺少id，已跳过。'
            )
            continue

        if line_name is None or not str(line_name).strip():
            print(
                f'警告：线路{line_id}缺少name，已跳过。'
            )
            continue

        result.append({
            'id': str(line_id).strip(),
            'name': str(line_name).strip(),
            'company': str(company).strip()
        })

    if not result:
        raise ValueError(
            'lines.json中没有有效线路。'
        )

    return result


def format_line_name(line):
    """返回JSON中配置的线路名称。"""

    return line['name']


def format_plan_run_time(plan_run_time, scan_date):
    """将planRunTime转换为Excel日期时间。"""

    if not plan_run_time:
        return None

    plan_run_time = str(plan_run_time).strip()

    for fmt in (
        '%Y-%m-%d %H:%M',
        '%Y-%m-%d %H:%M:%S'
    ):
        try:
            return datetime.strptime(
                f'{scan_date} {plan_run_time}',
                fmt
            )
        except ValueError:
            continue

    print(
        f'[{datetime.now().isoformat()}] '
        f'无法解析planRunTime：{plan_run_time}'
    )

    return None


def get_one_line_vehicle_info(line, mark, scan_date):
    """获取一条线路一个方向的车辆原始数据。"""

    line_id = line['id']
    line_name = format_line_name(line)

    direction = (
        '上行'
        if mark == 0
        else '下行'
    )

    session = get_session()

    try:
        response = session.get(
            API_URL.format(line_id, mark),
            timeout=REQUEST_TIMEOUT
        )

        response.raise_for_status()
        data = response.json()

    except Exception as e:
        print(
            f'[{datetime.now().isoformat()}] '
            f'{line_name} {direction}请求失败：{e}'
        )
        return []

    line_pipe_beans = data.get('linePipeBeans')

    if not isinstance(line_pipe_beans, list):
        return []

    result = []

    for item in line_pipe_beans:

        if not isinstance(item, dict):
            continue

        vehicle_id = item.get('vehicleId')
        vehicle_code = item.get('vehicleCode')
        plan_run_time = item.get('planRunTime')
        driver_code = item.get('driverCode')

        if (
            vehicle_id is None
            and vehicle_code is None
            and plan_run_time is None
            and driver_code is None
        ):
            continue

        departure_time = format_plan_run_time(
            plan_run_time,
            scan_date
        )

        result.append({
            'source_line': line,
            'direction': direction,
            'vehicle_id': vehicle_id,
            'vehicle_code': vehicle_code,
            'departure_time': departure_time,
            'driver_code': driver_code
        })

    return result


def process_group_data(raw_records, group_lines_map):
    """对交叉关联组的所有原始数据进行重新归位与去重清洗。"""

    cleaned_records = []
    seen_keys = set()

    for item in raw_records:
        vehicle_id = item['vehicle_id']
        if vehicle_id is None:
            continue

        vid_str = str(vehicle_id).strip()
        if not vid_str.isdigit():
            continue

        target_line = None

        # 核心判定规律：4位纯数字->镇江公交；6位纯数字->润港客运
        if len(vid_str) == 4:
            for lid, line_obj in group_lines_map.items():
                if lid not in ('971001', '971021', '971101'):
                    target_line = line_obj
                    break
        elif len(vid_str) == 6:
            for lid, line_obj in group_lines_map.items():
                if lid in ('971001', '971021', '971101'):
                    target_line = line_obj
                    break

        if not target_line:
            continue

        line_name = format_line_name(target_line)
        direction = item['direction']
        v_code = item['vehicle_code']
        dep_time = item['departure_time']
        d_code = item['driver_code']
        company = target_line.get('company', '')

        # 组内去重 Key（按前 6 位，防止同一条数据在两边 API 重复出现）
        dedup_key = (
            line_name,
            direction,
            str(vehicle_id).strip(),
            str(v_code).strip() if v_code is not None else None,
            dep_time,
            str(d_code).strip() if d_code is not None else None
        )

        if dedup_key in seen_keys:
            continue

        seen_keys.add(dedup_key)

        cleaned_records.append((
            line_name,
            direction,
            vehicle_id,
            v_code,
            dep_time,
            d_code,
            company
        ))

    return cleaned_records


def get_vehicle_data(lines):
    """并发获取所有线路上下行数据（自动交叉汇总清洗关联线路）。"""

    scan_date = datetime.now().strftime(
        '%Y-%m-%d'
    )

    lines_by_id = {str(l['id']): l for l in lines}

    normal_tasks = []
    group_tasks_map = {}

    for line in lines:
        lid = str(line['id'])
        if lid in LINE_GROUP_MAP:
            g_info = LINE_GROUP_MAP[lid]
            g_key = f"{g_info['zj_id']}_{g_info['rg_id']}"
            if g_key not in group_tasks_map:
                group_tasks_map[g_key] = []
            group_tasks_map[g_key].append((line, 0))
            group_tasks_map[g_key].append((line, 1))
        else:
            normal_tasks.append((line, 0))
            normal_tasks.append((line, 1))

    total_tasks = len(normal_tasks) + sum(len(v) for v in group_tasks_map.values())
    completed = 0

    print()
    print('=' * 60)
    print(
        f'[{datetime.now().isoformat()}] 开始扫描'
    )
    print(
        f'线路：{len(lines)}条（包含 {len(group_tasks_map)} 组关联交叉线路）'
    )
    print(
        f'请求：{total_tasks}个'
    )
    print(
        f'并发：{MAX_WORKERS}线程'
    )
    print('=' * 60)

    task_results_by_group = {g_key: [] for g_key in group_tasks_map}
    normal_cleaned_results = []

    with ThreadPoolExecutor(
        max_workers=MAX_WORKERS
    ) as executor:

        future_map = {}

        for line, mark in normal_tasks:
            f = executor.submit(get_one_line_vehicle_info, line, mark, scan_date)
            future_map[f] = ('normal', line, mark)

        for g_key, tasks in group_tasks_map.items():
            for line, mark in tasks:
                f = executor.submit(get_one_line_vehicle_info, line, mark, scan_date)
                future_map[f] = ('group', g_key, line, mark)

        for future in as_completed(future_map):
            completed += 1
            info = future_map[future]
            task_type = info[0]

            try:
                raw_data = future.result()

                if task_type == 'normal':
                    _, line, mark = info
                    line_name = format_line_name(line)
                    direction = '上行' if mark == 0 else '下行'
                    company = line.get('company', '')

                    for item in raw_data:
                        normal_cleaned_results.append((
                            line_name,
                            item['direction'],
                            item['vehicle_id'],
                            item['vehicle_code'],
                            item['departure_time'],
                            item['driver_code'],
                            company
                        ))

                    print(
                        f'[{completed}/{total_tasks}] '
                        f'{line_name} {direction}：'
                        f'{len(raw_data)}条'
                    )

                elif task_type == 'group':
                    _, g_key, line, mark = info
                    line_name = format_line_name(line)
                    direction = '上行' if mark == 0 else '下行'

                    task_results_by_group[g_key].extend(raw_data)

                    print(
                        f'[{completed}/{total_tasks}] '
                        f'(关联组) {line_name} {direction}：'
                        f'{len(raw_data)}条'
                    )

            except Exception as e:
                print(
                    f'[{completed}/{total_tasks}] '
                    f'处理失败：{e}'
                )

    # 对关联组的数据进行归位与去重
    final_group_results = []
    for g_key, raw_group_data in task_results_by_group.items():
        zj_id, rg_id = g_key.split('_')
        g_lines_map = {}
        if zj_id in lines_by_id:
            g_lines_map[zj_id] = lines_by_id[zj_id]
        if rg_id in lines_by_id:
            g_lines_map[rg_id] = lines_by_id[rg_id]

        cleaned_group_data = process_group_data(raw_group_data, g_lines_map)
        final_group_results.extend(cleaned_group_data)

    result = normal_cleaned_results + final_group_results

    print(
        f'[{datetime.now().isoformat()}] '
        f'扫描结束，共获取{len(result)}条。'
    )

    return result


def normalize_datetime(value):
    """统一Excel日期时间格式。"""

    if value is None:
        return None

    if isinstance(value, datetime):
        return value.replace(
            second=0,
            microsecond=0
        )

    value = str(value).strip()

    for fmt in (
        '%Y-%m-%d %H:%M',
        '%Y-%m-%d %H:%M:%S'
    ):
        try:
            return datetime.strptime(value, fmt)
        except ValueError:
            continue

    return value


def make_record_key(record):
    """生成历史记录唯一键（仍使用前6列生成唯一键，保持与旧数据一致）。"""

    departure_time = normalize_datetime(
        record[4]
    )

    values = [
        record[0],
        record[1],
        record[2],
        record[3],
        departure_time,
        record[5]
    ]

    return tuple(
        None if value is None
        else str(value).strip()
        if not isinstance(value, datetime)
        else value
        for value in values
    )


def setup_excel_format(ws):
    """设置Excel筛选、冻结和列宽。"""

    ws.auto_filter.ref = (
        f'A1:G{max(ws.max_row, 1)}'
    )

    ws.freeze_panes = 'A2'

    widths = {
        'A': 20,
        'B': 10,
        'C': 15,
        'D': 15,
        'E': 22,
        'F': 12,
        'G': 15
    }

    for column, width in widths.items():
        ws.column_dimensions[column].width = width

    for row in range(2, ws.max_row + 1):
        cell = ws.cell(
            row=row,
            column=5
        )

        if isinstance(cell.value, datetime):
            cell.number_format = (
                'yyyy-mm-dd hh:mm'
            )


def create_excel(filepath):
    """创建新的历史数据Excel。"""

    wb = Workbook()
    ws = wb.active
    ws.title = '车辆信息'

    for column, header in enumerate(
        HEADERS,
        start=1
    ):
        ws.cell(
            row=1,
            column=column,
            value=header
        )

    setup_excel_format(ws)

    wb.save(filepath)
    wb.close()


def append_unique_to_excel(vehicle_list, filepath, lines_config):
    """追加新记录，并自动补全旧数据中缺失的运营公司"""

    if not os.path.exists(filepath):
        create_excel(filepath)

    wb = load_workbook(filepath)
    ws = wb.active

    # 1. 确保表头更新为 7 列
    for column, header in enumerate(
        HEADERS,
        start=1
    ):
        ws.cell(
            row=1,
            column=column,
            value=header
        )

    # 建立“线路名称 -> 运营公司”的查找词典，用于为旧表格数据补全 company
    name_to_company = {l['name']: l['company'] for l in lines_config}

    existing = set()

    # 2. 读取并检查旧数据，给缺失 company 的历史行自动补充
    for r_idx in range(2, ws.max_row + 1):
        row_cells = [ws.cell(row=r_idx, column=c) for c in range(1, 8)]
        row_vals = [cell.value for cell in row_cells]

        if all(value is None for value in row_vals):
            continue

        # 前6位存入唯一键集合，用于新记录去重
        existing.add(
            make_record_key(row_vals[:6])
        )

        # 补全既有数据中空的“运营公司”列（第7列）
        line_name = row_vals[0]
        company_val = row_vals[6]

        if (company_val is None or str(company_val).strip() == '') and line_name:
            if line_name in name_to_company:
                ws.cell(
                    row=r_idx,
                    column=7,
                    value=name_to_company[line_name]
                )

    # 3. 追加新获取的数据
    to_append = []

    for item in vehicle_list:

        key = make_record_key(item[:6])

        if key in existing:
            continue

        existing.add(key)
        to_append.append(item)

    for item in to_append:

        row = ws.max_row + 1

        for column, value in enumerate(
            item,
            start=1
        ):
            ws.cell(
                row=row,
                column=column,
                value=value
            )

        if isinstance(item[4], datetime):
            ws.cell(
                row=row,
                column=5
            ).number_format = (
                'yyyy-mm-dd hh:mm'
            )

    setup_excel_format(ws)

    wb.save(filepath)
    wb.close()

    return len(to_append)


def run_one_scan(lines):
    """执行一轮扫描并保存数据。"""

    vehicle_data = get_vehicle_data(lines)

    added = append_unique_to_excel(
        vehicle_data,
        EXCEL_FILE,
        lines
    )

    print(
        f'[{datetime.now().isoformat()}] '
        f'本轮获取{len(vehicle_data)}条，'
        f'新增{added}条。'
    )

    return added


if __name__ == '__main__':
    print(
        'zjgj_Fleet&Plate_Collector v3.5'
    )
    print('=' * 60)

    try:
        lines = load_lines()

    except Exception as e:

        print(
            f'线路配置读取失败：{e}'
        )

        input('按Enter键退出……')
        raise SystemExit(1)

    print(
        f'线路配置：{LINE_CONFIG_FILE}'
    )

    print(
        f'线路数量：{len(lines)}'
    )

    print(
        f'扫描间隔：{INTERVAL_SECONDS // 60}分钟'
    )

    print(
        f'并发线程：{MAX_WORKERS}'
    )

    print(
        f'历史数据：{EXCEL_FILE}'
    )

    try:
        run_one_scan(lines)

    except Exception as e:

        print(
            f'首次扫描失败：{e}'
        )

    try:

        while True:

            print()
            print(
                f'[{datetime.now().isoformat()}] '
                f'等待10分钟后进行下一轮扫描……'
            )

            time.sleep(
                INTERVAL_SECONDS
            )

            try:
                run_one_scan(lines)

            except Exception as e:

                print(
                    f'周期扫描失败：{e}'
                )

    except KeyboardInterrupt:

        print()
        print(
            f'[{datetime.now().isoformat()}] '
            f'程序已退出。'
        )
